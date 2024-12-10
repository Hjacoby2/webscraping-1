import openpyxl as xl
from openpyxl.styles import Font

wb = xl.Workbook()
ws = wb.active
ws.title = 'First Sheet'
wb.create_sheet(index=1, title='Second Sheet')

ws['A1'] = 'Invoice'
ws['A1'].font = Font(name='Times New Roman', size=24, bold=True)
fontObj = Font(name='Times New Roman', size=24, bold=True, italic=False)
ws['A1'].font = fontObj

ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = 'Alignment'

ws['B2'] = 450
ws['B3'] = 225.50
ws['B4'] = 150


ws['A8'] = 'Total'
ws['A8'].font = fontObj
ws.merge_cells('A8:B8')



ws.column_dimensions['A'].width = 25

ws.merge_cells('A1:B1')

ws['B8'] = '=SUM(B2:B7)'

write_sheet = wb['Second Sheet']

read_wb = xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb['ProduceReport']

write_sheet['A1'] = 'Produce'
write_sheet['B1'] = 'Cost Per Pound'
write_sheet['C1'] = 'Amt Sold'
write_sheet['D1'] = 'Total'

rowcounter = 2

for row in read_ws.iter_rows(min_row=2, values_only=True):
    name = row[0]
    cost = row[1]
    amt_sold = row[2]
    total = row[3]

    write_sheet.cell(row=rowcounter, column=1).value = name
    write_sheet.cell(row=rowcounter, column=2).value = cost
    write_sheet.cell(row=rowcounter, column=3).value = amt_sold
    write_sheet.cell(row=rowcounter, column=4).value = total

    rowcounter += 1

summary_row = rowcounter + 1
write_sheet.merge_cells(f'B{summary_row}:C{summary_row}')
write_sheet['B' + str(summary_row)] = 'Total'
write_sheet['B' + str(summary_row)].font = fontObj

write_sheet['C' + str(summary_row)] = f'=AVERAGE(C2:C{rowcounter-1})'
write_sheet['D' + str(summary_row)] = f'=SUM(D2:D{rowcounter-1})'

write_sheet.column_dimensions['A'].width = 16
write_sheet.column_dimensions['B'].width = 15
write_sheet.column_dimensions['C'].width = 15
write_sheet.column_dimensions['D'].width = 15

for cell in write_sheet["C"]:
    cell.number_format = '#,##0'

for cell in write_sheet['D']:
    cell.number_format = u'"$"#,##0.00'

wb.save('PythontoExcel.xlsx')
