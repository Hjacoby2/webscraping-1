from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl


webpage = 'https://www.boxofficemojo.com/year/2024/'

page = urlopen(webpage)			
soup = BeautifulSoup(page, 'html.parser')


title = soup.title
print(title.text)

movie_rows = soup.find_all('tr')


wb = xl.Workbook()
ws = wb.active
ws.title = 'Box Office Report'


ws['A1'] = 'Movie Title'
ws['B1'] = 'Gross Returns'
ws['C1'] = 'No. of Theaters'
ws['D1'] = 'Average/Theater'


for x in range(1, 6):  
    td = movie_rows[x].find_all('td')

    title = td[1].text.strip()
    gross = td[5].text.replace("$", "").replace(",", "").strip()
    theater = int(td[6].text.replace(",", "").strip())

    avg = round(int(gross) / theater, 2)

    ws['A' + str(x+1)].value = title
    ws['B' + str(x+1)].value = gross
    ws['C' + str(x+1)].value = theater
    ws['D' + str(x+1)].value = avg  

wb.save('BoxOfficeReport.xlsx')
